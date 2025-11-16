import re
import ast
import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed
import os

class ArabicProcessor:
    def __init__(self, symbols_path):
        self.arabic_symbols = self.load_arabic_symbols(symbols_path)
        self.compiled_patterns = {}

    def load_arabic_symbols(self, file_path):
        with open(file_path, "r", encoding="utf-8") as file:
            return ast.literal_eval(file.read())

    def compile_weights_to_regex(self, weights):
        for weight in weights:
            pattern = "".join(self.arabic_symbols.get(letter, letter) for letter in weight)
            self.compiled_patterns[weight] = re.compile(pattern)

def unify_hamzas(word):
    hamzas = 'ئءؤأإآ'
    unified_word = ''.join('ء' if c in hamzas else c for c in word)
    return unified_word

def remove_diacritics_and_unify_hamzas(word):
    diacritics = 'َُِْٰ'
    cleaned_word = ''.join(c for c in word if c not in diacritics)
    return unify_hamzas(cleaned_word)

def find_equivalent(left_word, right_word):
    left_word_clean = remove_diacritics_and_unify_hamzas(left_word)
    right_word_clean = remove_diacritics_and_unify_hamzas(right_word)
    
    mapping = {char: '' for char in 'فعل'}
    for idx, char in enumerate(right_word_clean):
        if char in mapping and idx < len(left_word_clean):
            mapping[char] = left_word_clean[idx]
            
    return ''.join(mapping[char] for char in 'فعل')


class FileManager:
    def __init__(self, processor):
        self.processor = processor

    def read_sarf_weights(self, file_path):
        special_chars = set('أإآؤئستمنهيءوج')
        with open(file_path, 'r', encoding='utf-8') as file:
            weights = [line.strip() for line in file]
        
        priority_weights = [w for w in weights if any(c in w for c in special_chars)]
        secondary_weights = [w for w in weights if w not in priority_weights]

        return priority_weights, secondary_weights

    def process_words_from_excel(self, excel_file_path, weights_file_path, symbols_file_path, max_workers=None):
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active
        
        # تحديث تسميات الأعمدة
        ws.cell(row=1, column=3, value="الأوزان")  # تسمية العمود C
        ws.cell(row=1, column=4, value="الجذور2")  # تسمية العمود D
        if 'المقارنة' not in [cell.value for cell in ws[1]]:
            ws.cell(row=1, column=5, value="المقارنة")  # إضافة عنوان العمود E إذا لم يكن موجودا
      
        priority_weights, secondary_weights = self.read_sarf_weights(weights_file_path)
        all_weights = priority_weights + secondary_weights

        self.processor.compile_weights_to_regex(all_weights)

        # قراءة جميع الكلمات والجمليات يدويًا
        words_data = []
        for row in range(2, ws.max_row + 1):
            root_cell = ws.cell(row=row, column=1)  
            word_cell = ws.cell(row=row, column=2)  
            root_manual = root_cell.value.strip() if root_cell.value else ""
            word = word_cell.value.strip() if word_cell.value else ""
            
            if not word: 
                continue

            words_data.append((row, root_manual, word))

        def process_single_word(data):
            row, root_manual, word = data
            matched = False
            for weight in all_weights:
                pattern = self.processor.compiled_patterns[weight]
                if pattern.fullmatch(word):
                    equivalent_root = find_equivalent(word, weight)
                    match_status = "مطابقة" if equivalent_root == root_manual else "غير مطابقة"
                    return (row, weight, equivalent_root, match_status, matched := (match_status == "مطابقة"))
            return (row, "لا يوجد", "", "لا يوجد وزن ملائم", False)

        results = []
        # تحديد عدد الأنوية إذا لم يتم تحديده
        if max_workers is None:
            max_workers = os.cpu_count() or 4  # استخدام عدد الأنوية الافتراضي أو 4 إذا كان غير معروف

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_data = {executor.submit(process_single_word, data): data for data in words_data}
            for future in as_completed(future_to_data):
                result = future.result()
                results.append(result)

        # تحديث ملف Excel بناءً على النتائج
        for result in results:
            row, weight, equivalent_root, match_status, matched = result
            ws.cell(row=row, column=3, value=weight) 
            ws.cell(row=row, column=4, value=equivalent_root)  
            ws.cell(row=row, column=5, value=match_status)

        wb.save(excel_file_path)


if __name__ == "__main__":
    excel_file_path = "الألفاظ.xlsx"
    weights_file_path = "الأوزان++++.txt"
    symbols_file_path = "الخريطة.txt"
    
    arabic_processor = ArabicProcessor(symbols_file_path)
    file_manager = FileManager(arabic_processor)
    
    # يمكنك تحديد عدد الأنوية هنا، مثلاً 8
    file_manager.process_words_from_excel(excel_file_path, weights_file_path, symbols_file_path, max_workers=8)
