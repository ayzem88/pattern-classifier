import re
import ast
import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed
import os

# كلاس لمعالجة الرموز العربية وتحويل الأوزان إلى أنماط ريجيكس
class ArabicProcessor:
    def __init__(self, symbols_path):
        self.arabic_symbols = self.load_arabic_symbols(symbols_path)
        self.compiled_patterns = {}

    # دالة لقراءة وتحميل رموز العربية من ملف
    def load_arabic_symbols(self, file_path):
        with open(file_path, "r", encoding="utf-8") as file:
            return ast.literal_eval(file.read())

    # دالة لتحويل الأوزان إلى أنماط ريجيكس وتخزينها
    def compile_weights_to_regex(self, weights):
        for weight in weights:
            # تحويل الوزن إلى نمط ريجيكس باستخدام الرموز العربية مع بداية ونهاية السلسلة
            pattern = "^" + "".join(self.arabic_symbols.get(letter, letter) for letter in weight) + "$"
            self.compiled_patterns[weight] = re.compile(pattern)

# كلاس لمعالجة النصوص العربية (توحيد الهمزات، إزالة التشكيل، إيجاد الجذر المكافئ)
class ArabicTextProcessor:
    def __init__(self):
        pass

    # دالة لتوحيد جميع أشكال الهمزات إلى "ء"
    def unify_hamzas(self, word):
        hamzas = 'ئءؤأإآ'
        unified_word = ''.join('ء' if c in hamzas else c for c in word)
        return unified_word

    # دالة لإزالة التشكيل وتوحيد الهمزات في الكلمة
    def remove_diacritics_and_unify_hamzas(self, word):
        diacritics = 'َُِْٰ'
        cleaned_word = ''.join(c for c in word if c not in diacritics)
        return self.unify_hamzas(cleaned_word)

    # دالة لإيجاد الجذر المكافئ بين كلمتين
    def find_equivalent(self, left_word, right_word):
        left_word_clean = self.remove_diacritics_and_unify_hamzas(left_word)
        right_word_clean = self.remove_diacritics_and_unify_hamzas(right_word)
        
        # تحديد طول الجذر بناءً على الجذر اليدوي (من العمود الأول)
        root_length = len(left_word_clean)
        if root_length not in [3, 4]:
            # يمكنك إضافة دعم لأطوال جذور أخرى إذا لزم الأمر
            raise ValueError("طول الجذر غير مدعوم. يجب أن يكون 3 أو 4 حروف.")
        
        # تكييف الحروف بناءً على طول الجذر
        mapping = {char: '' for char in left_word_clean}
        for idx, char in enumerate(right_word_clean):
            if char in mapping and idx < len(left_word_clean):
                mapping[char] = left_word_clean[idx]
                
        return ''.join(mapping.get(char, '') for char in left_word_clean)

# كلاس للتعامل مع قلب الأوزان (تبديل وزن بوزن آخر)
class WeightReverser:
    def __init__(self, reversal_file_path):
        self.reversal_mapping = self.load_reversal_mapping(reversal_file_path)
    
    # دالة لقراءة خريطة قلب الأوزان من ملف
    def load_reversal_mapping(self, file_path):
        mapping = {}
        with open(file_path, "r", encoding="utf-8") as file:
            for line in file:
                # تجاهل الأسطر الفارغة والأسطر التي لا تحتوي على '='
                if "=" in line:
                    key, value = line.split("=", 1)
                    key = key.strip()
                    value = value.strip()
                    mapping[key] = value
        return mapping

    # دالة للحصول على الوزن البديل إذا كان موجوداً
    def get_replacement(self, weight):
        return self.reversal_mapping.get(weight, None)

# كلاس لإدارة أولوية الجذور (ترتيب الكلمات بناءً على طول الجذر)
class RootPriorityManager:
    def __init__(self):
        pass

    # دالة لترتيب بيانات الكلمات بناءً على طول الجذر (الأطول أولاً)
    def prioritize_words_data(self, words_data):
        words_data_sorted = sorted(words_data, key=lambda x: len(x[1]), reverse=True)
        return words_data_sorted

# كلاس لمعالجة كلمة واحدة (تحديد الوزن والجذر المكافئ وحالة المطابقة)
class WordProcessor:
    def __init__(self, arabic_processor, weight_reverser):
        self.arabic_processor = arabic_processor
        self.weight_reverser = weight_reverser
        self.text_processor = ArabicTextProcessor()

    # دالة لمعالجة كلمة واحدة وإيجاد الوزن والجذر وحالة المطابقة
    def process_single_word(self, data, all_weights, root_manual):
        row, word = data
        matching_weights_with_root = []
        matching_weights_without_root = []
        root_length = len(root_manual)
        for weight in all_weights:
            pattern = self.arabic_processor.compiled_patterns[weight]
            if pattern.fullmatch(word):
                try:
                    equivalent_root = self.text_processor.find_equivalent(word, weight)
                except ValueError:
                    # إذا كان طول الجذر غير مدعوم، اعتبره غير مطابق
                    equivalent_root = ""
                if equivalent_root == root_manual:
                    matching_weights_with_root.append((weight, equivalent_root, "مطابقة"))
                else:
                    matching_weights_without_root.append((weight, equivalent_root, "غير مطابقة"))
        
        # محاولة إيجاد أوزان مع تطابق الجذر
        if matching_weights_with_root:
            weight, equivalent_root, status = matching_weights_with_root[0]
            replacement = self.weight_reverser.get_replacement(weight)
            if replacement:
                weight = replacement
                status = "مقلوب"
            return (row, weight, equivalent_root, status)
        
        # إذا لم يوجد أوزان مع تطابق الجذر، اختر الوزن الذي يتطابق مع الكلمة واعتبره غير مطابق
        if matching_weights_without_root:
            weight, equivalent_root, status = matching_weights_without_root[0]
            replacement = self.weight_reverser.get_replacement(weight)
            if replacement:
                weight = replacement
                status = "مقلوب"
            else:
                status = "وزن تقريبيّ"
            return (row, weight, equivalent_root, status)
        
        # إذا لم يوجد وزن يتطابق مع الكلمة
        return (row, "لا يوجد", "", "غير معروف")

# كلاس لإدارة قراءة الملفات ومعالجة الكلمات وتخزين النتائج في ملف إكسل
class FileManager:
    def __init__(self, processor, reverser):
        self.processor = processor
        self.reverser = reverser
        self.priority_manager = RootPriorityManager()
        self.word_processor = WordProcessor(self.processor, self.reverser)

    # دالة لقراءة الأوزان الصرفية من ملف
    def read_sarf_weights(self, file_path):
        with open(file_path, 'r', encoding='utf-8') as file:
            weights = [line.strip() for line in file if line.strip()]
        return weights

    # دالة لمعالجة الكلمات من ملف إكسل وتطبيق الأوزان عليها وتخزين النتائج
    def process_words_from_excel(self, excel_file_path, weights_file_path, symbols_file_path, reversal_file_path, max_workers=None):
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active
        
        # إعداد عناوين الأعمدة في ملف الإكسل
        ws.cell(row=1, column=3, value="الأوزان")  # تسمية العمود C
        ws.cell(row=1, column=4, value="الجذور2")  # تسمية العمود D
        if 'المقارنة' not in [cell.value for cell in ws[1]]:
            ws.cell(row=1, column=5, value="المقارنة")  # إضافة عنوان العمود E إذا لم يكن موجودا

        # قراءة الأوزان الصرفية
        all_weights = self.read_sarf_weights(weights_file_path)

        # تحويل الأوزان إلى أنماط ريجيكس
        self.processor.compile_weights_to_regex(all_weights)

        words_data = []
        for row in range(2, ws.max_row + 1):
            root_cell = ws.cell(row=row, column=1)  
            word_cell = ws.cell(row=row, column=2)  
            root_manual = root_cell.value.strip() if root_cell.value else ""
            word = word_cell.value.strip() if word_cell.value else ""
            
            if not word: 
                continue

            words_data.append((row, root_manual, word))

        # ترتيب الكلمات بناءً على طول الجذر (الأطول أولاً)
        words_data = self.priority_manager.prioritize_words_data(words_data)

        results = []

        if max_workers is None:
            max_workers = os.cpu_count() or 4  # استخدام عدد الأنوية الافتراضي أو 4 إذا كان غير معروف

        # دالة لمعالجة البيانات لكلمة واحدة
        def process_data(data):
            row, root_manual, word = data
            return self.word_processor.process_single_word((row, word), all_weights, root_manual)

        # استخدام ThreadPoolExecutor لمعالجة الكلمات بشكل متعدد الأنوية مع تتبع التقدم
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_data = {executor.submit(process_data, data): data for data in words_data}
            completed_count = 0
            for future in as_completed(future_to_data):
                result = future.result()
                results.append(result)
                completed_count += 1
                if completed_count % 1000 == 0:
                    print(f"انتهى العمل من {completed_count}، جاري معالجة من {completed_count+1} إلى {completed_count+1000}")

        # تخزين النتائج في ملف الإكسل
        for result in results:
            row, weight, equivalent_root, match_status = result
            ws.cell(row=row, column=3, value=weight) 
            ws.cell(row=row, column=4, value=equivalent_root)  
            ws.cell(row=row, column=5, value=match_status)

        wb.save(excel_file_path)

# نقطة البداية لتنفيذ البرنامج
if __name__ == "__main__":
    excel_file_path = "الألفاظ_كامل.xlsx"
    weights_file_path = "الأوزان+.txt"
    symbols_file_path = "الخريطة.txt"
    reversal_file_path = "قلب الأوزان.txt"  # مسار ملف قلب الأوزان
    
    # إنشاء كائنات الكلاسات المختلفة
    arabic_processor = ArabicProcessor(symbols_file_path)
    weight_reverser = WeightReverser(reversal_file_path)
    file_manager = FileManager(arabic_processor, weight_reverser)
    
    # يمكنك تحديد عدد الأنوية هنا، مثلاً 8
    file_manager.process_words_from_excel(
        excel_file_path, 
        weights_file_path, 
        symbols_file_path, 
        reversal_file_path,  # تمرير مسار ملف القلب
        max_workers=8
    )
